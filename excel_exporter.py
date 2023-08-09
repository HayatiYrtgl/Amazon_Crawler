import time
from os import remove
from openpyxl import Workbook

# excel file woekbook

excel_w = Workbook()

excel_cll = excel_w.active

# excel name with time

created_time = time.strftime('%H.%M.%S')

# excel configuration

excel_cll.cell(row=1, column=1, value="Ürün adı")
excel_cll.cell(row=1, column=2, value="Ürün Fıyatı")
excel_cll.cell(row=1, column=3, value="Yıldız")
excel_cll.cell(row=1, column=4, value="Ürün Linki")
excel_cll.cell(row=1, column=5, value="Görüntülenme")
excel_cll.cell(row=1, column=6, value="Shipping ücreti")
excel_cll.cell(row=1, column=7, value="Stock durumu")
excel_cll.cell(row=1, column=8, value="Gönderen")
excel_cll.cell(row=1, column=9, value="Satıcı")

def import_excel_file_main(product_number_plus_one, v1, v2, v3, v4):

    # arguments for excel rows
    excel_cll.cell(row=product_number_plus_one + 1, column=1, value=v1)
    excel_cll.cell(row=product_number_plus_one + 1, column=2, value=v2)
    excel_cll.cell(row=product_number_plus_one + 1, column=3, value=v3)
    excel_cll.cell(row=product_number_plus_one + 1, column=4, value=v4)

    excel_w.save(f"amazon-{created_time}.xlsx")

def import_excel_file_detail(product_number_plus_one, v1, v2, v3, v4, v5):

    excel_cll.cell(row=product_number_plus_one + 1, column=5, value=v1)
    excel_cll.cell(row=product_number_plus_one + 1, column=6, value=v2)
    excel_cll.cell(row=product_number_plus_one + 1, column=7, value=v3)
    excel_cll.cell(row=product_number_plus_one + 1, column=8, value=v4)
    excel_cll.cell(row=product_number_plus_one + 1, column=9, value=v5)

    excel_w.save(f"amazon-{created_time}.xlsx")